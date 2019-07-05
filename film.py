
import os
import re
import time

import requests
import xlsxwriter
from bs4 import BeautifulSoup
from openpyxl import load_workbook
from openpyxl.drawing.image import Image
from PIL import Image as PIL_IMAGE

HEADER = {'User-Agent':'Mozilla/5.0 (Windows NT 10.0; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/58.0.3029.110 Safari/537.36 SE 2.X MetaSr 1.0'}
START_PAGE = 1
END_PAGE = 41
XLSX_NAME = "恐怖电影.xlsx"
XLSX_SHEET_NAME = "恐怖电影"
FOLDER_NAME = "电影图片"
FILM_LENGTH = 0

def makeWorkbook():
    workbook = xlsxwriter.Workbook(XLSX_NAME)
    worksheet = workbook.add_worksheet(XLSX_SHEET_NAME)
    # 图片
    worksheet.set_column('A:A', 40)
    # 标题
    worksheet.set_column('B:B', 100)
    # 页数
    worksheet.set_column('C:C', 10)
    # 电影地址
    worksheet.set_column('D:D', 50)
    # 图片URL
    worksheet.set_column('E:E', 50)
    return workbook

def readAndDownloadPictures():
    workbook = load_workbook(XLSX_NAME)
    sheet = workbook[XLSX_SHEET_NAME]
    rows = sheet.max_row
    # 电影序号以0开始
    for i in range(1, rows + 1):
        imgUrl = sheet.cell(row = i, column = 5).value
        if imgUrl != None:
            downloadPicture(imgUrl, i - 1)

def insertPictures():
    print("开始插入图片")
    workbook = load_workbook(XLSX_NAME)
    sheet = workbook[XLSX_SHEET_NAME]
    
    for currdir, subdir, fileList in os.walk(FOLDER_NAME):
        for filename in fileList:
            try:
                imgPath = os.path.join(currdir, filename)
                num = re.findall(r'(.+?)\.', filename)[0]
                num = int(num) + 1
                area = 'A' + str(num)
                img = Image(imgPath)
                sheet.column_dimensions['A'].width = 33
                sheet.row_dimensions[num].height = 285
                sheet.add_image(img, area)
            except OSError:
                print(filename + "文件损坏")
    workbook.save(XLSX_NAME)
    print("插入图片完成")

def getHtml(workbook, filmLen):
    titleFormat = workbook.add_format({'font_size':24})
    worksheet = workbook.get_worksheet_by_name(XLSX_SHEET_NAME)
    # 记录电影序号
    for i in range(START_PAGE,END_PAGE):
        url = "http://rarbt.cc/index.php/category/index/id/9/p/{0}.html".format(i)
        request = requests.get(url, headers = HEADER)
        soup = BeautifulSoup(request.text, 'html.parser')
        filmList = soup.find_all(attrs={'class':'item cl'})

        pageNum = "第{0}页".format(i)
        print(pageNum + " 开始")

        for film in filmList:
            fileUrl = film.find('a')
            imgInfo = film.find('img')
            title = imgInfo['alt']
            imgUrl = imgInfo['src']
            filmUrl = "http://rarbt.cc" + fileUrl['href']
            if imgUrl == '':
                imgUrl = ''
            elif imgUrl.find('http://') == -1:
                imgUrl = "http://rarbt.cc" + imgUrl
            worksheet.write(filmLen, 1, title, titleFormat)
            worksheet.write(filmLen, 2, pageNum)
            worksheet.write(filmLen, 3, filmUrl)
            worksheet.write(filmLen, 4, imgUrl)
            filmLen += 1
        
        print(pageNum + " 完成")

def imgResize(path):
    width = 270
    height = 380
    print("开始图片缩放")
    for currdir, subdir, fileList in os.walk(path):
        for filename in fileList:
            imgPath = os.path.join(currdir, filename)
            try:
                img = PIL_IMAGE.open(imgPath)
                img = img.resize((width, height), PIL_IMAGE.ANTIALIAS)
                img.save(imgPath)
            except OSError:
                print(filename + "文件损坏")
    print("图片缩放完成")

def downloadPicture(url, fileName):
    if not os.path.exists(FOLDER_NAME):
        os.makedirs(FOLDER_NAME)
    try:
        request = requests.get(url)
        fileName = str(fileName) + '.jpg'
        print(fileName)
        with open(os.path.join(FOLDER_NAME, fileName), 'wb') as file:
            file.write(request.content)
    except BaseException:
        print(str(fileName) + "下载失败")
    
def main():
    startTime = time.time()
    workbook = makeWorkbook()
    # 获取电影标题, 图片url
    getHtml(workbook, FILM_LENGTH)
    workbook.close()
    # 读取并下载图片url
    readAndDownloadPictures()
    # 缩放图片
    imgResize(FOLDER_NAME)
    # 插入图片
    insertPictures()

    endTime = time.time()
    useTime = "查询完毕, 总用时:%f"%(endTime - startTime)
    print(useTime)

if __name__ == "__main__":
    main()
